"""
Ingestion catalogue Boss → embeddings → PostgreSQL.

Usage:
  cd backend
  python -m scripts.ingest_catalog --xlsx ../articles_industriels_1000.xlsx
  python -m scripts.ingest_catalog --csv catalog.csv
  python -m scripts.ingest_catalog --xlsx ../articles_industriels_1000.xlsx --parse-only
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

XLSX_COLUMN_ALIASES: dict[str, list[str]] = {
    "reference": ["référence", "reference"],
    "designation": ["désignation", "designation"],
    "conditionnement": [
        "qtité conditionnement",
        "qtite conditionnement",
        "quantité conditionnement",
        "qte conditionnement",
    ],
    "alliage": ["alliage"],
    "prix_boite": ["prix boîte (€)", "prix boite (€)", "prix boîte", "prix boite"],
    "en_stock": ["en stock"],
    "delai_livraison": ["délai livraison", "delai livraison"],
}


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _cell(row: tuple[Any, ...], index: int) -> Any:
    if index >= len(row):
        return None
    return row[index]


def _map_xlsx_headers(headers: list[Any]) -> dict[str, int]:
    normalized = {_normalize_header(header): idx for idx, header in enumerate(headers)}
    mapping: dict[str, int] = {}

    for field, aliases in XLSX_COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                mapping[field] = normalized[alias]
                break
        if field not in mapping:
            raise ValueError(
                f"Colonne « {field} » introuvable. En-têtes du fichier : {list(headers)}"
            )

    return mapping


def _parse_en_stock(value: Any) -> bool:
    if value is None:
        return False
    return str(value).strip().lower() in {"oui", "yes", "true", "1"}


def _build_description(*, conditionnement: int, en_stock: bool, delai_livraison: str) -> str:
    stock_label = "en stock" if en_stock else "sur commande"
    return (
        f"Boîte de {conditionnement} pièces — {stock_label} — "
        f"délai {delai_livraison}"
    )


def _build_metadata(
    *,
    conditionnement: int,
    en_stock: bool,
    delai_livraison: str,
) -> dict[str, Any]:
    return {
        "conditionnement": conditionnement,
        "en_stock": en_stock,
        "delai_livraison": delai_livraison,
        "prix_unite": "boite",
    }


def load_catalog_rows_from_xlsx(xlsx_path: Path) -> list[dict[str, Any]]:
    """Lit articles_industriels_1000.xlsx et retourne des lignes normalisées."""
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.active

    rows_iter = worksheet.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    column_map = _map_xlsx_headers(headers)

    catalog_rows: list[dict[str, Any]] = []

    for row in rows_iter:
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        reference = str(_cell(row, column_map["reference"])).strip()
        designation = str(_cell(row, column_map["designation"])).strip()
        if not reference or not designation:
            continue

        conditionnement = int(_cell(row, column_map["conditionnement"]))
        alliage = str(_cell(row, column_map["alliage"])).strip()
        prix_boite = float(_cell(row, column_map["prix_boite"]))
        en_stock = _parse_en_stock(_cell(row, column_map["en_stock"]))
        delai_livraison = str(_cell(row, column_map["delai_livraison"])).strip()

        catalog_rows.append(
            {
                "sku": reference,
                "name": designation,
                "category": alliage,
                "unit_price": prix_boite,
                "stock_quantity": conditionnement if en_stock else 0,
                "description": _build_description(
                    conditionnement=conditionnement,
                    en_stock=en_stock,
                    delai_livraison=delai_livraison,
                ),
                "metadata": _build_metadata(
                    conditionnement=conditionnement,
                    en_stock=en_stock,
                    delai_livraison=delai_livraison,
                ),
            }
        )

    workbook.close()
    return catalog_rows


def load_catalog_rows_from_csv(csv_path: Path) -> list[dict[str, Any]]:
    with csv_path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        return [
            {
                "sku": row["sku"].strip(),
                "name": row["name"].strip(),
                "description": row.get("description", "").strip(),
                "category": row.get("category", "").strip(),
                "unit_price": float(row["unit_price"]),
                "stock_quantity": int(row.get("stock_quantity") or 0),
                "metadata": json.loads(row["metadata"]) if row.get("metadata") else {},
            }
            for row in reader
        ]


async def ingest_row(pool, row: dict[str, Any]) -> None:
    from app.services.embeddings import generate_embedding
    from app.services.vector_search import _embedding_to_pgvector_literal

    text_for_embed = " | ".join(
        filter(
            None,
            [
                row["sku"],
                row["name"],
                row.get("description") or "",
                row.get("category") or "",
            ],
        )
    )
    vector = await generate_embedding(text_for_embed)
    literal = _embedding_to_pgvector_literal(vector)

    await pool.execute(
        """
        INSERT INTO catalog_items (
            sku, name, description, category, unit_price,
            stock_quantity, metadata, embedding
        )
        VALUES ($1, $2, $3, $4, $5, $6, $7::jsonb, $8::vector)
        ON CONFLICT (sku) DO UPDATE SET
            name = EXCLUDED.name,
            description = EXCLUDED.description,
            category = EXCLUDED.category,
            unit_price = EXCLUDED.unit_price,
            stock_quantity = EXCLUDED.stock_quantity,
            metadata = EXCLUDED.metadata,
            embedding = EXCLUDED.embedding,
            updated_at = NOW()
        """,
        row["sku"],
        row["name"],
        row.get("description") or None,
        row.get("category") or None,
        row["unit_price"],
        row.get("stock_quantity", 0),
        json.dumps(row.get("metadata") or {}),
        literal,
    )


async def ingest_catalog(
    rows: list[dict[str, Any]],
    *,
    batch_pause: float,
    start_at: int,
) -> None:
    from app.db import close_pool, get_pool

    pool = await get_pool()

    for index, row in enumerate(rows, start=1):
        if index < start_at:
            continue

        await ingest_row(pool, row)

        if index % 50 == 0 or index == len(rows):
            print(f"Ingestion {index}/{len(rows)} — {row['sku']} — {row['name'][:50]}")

        if batch_pause > 0:
            await asyncio.sleep(batch_pause)

    await close_pool()


def _print_parse_preview(rows: list[dict[str, Any]], limit: int = 3) -> None:
    print(f"OK — {len(rows)} articles lus depuis le fichier")
    for row in rows[:limit]:
        print(
            f"  - {row['sku']} | {row['name']} | {row['category']} | "
            f"{row['unit_price']} €/boîte | stock={row['stock_quantity']} | "
            f"{row['metadata']}"
        )
    if len(rows) > limit:
        print(f"  … et {len(rows) - limit} autres lignes")


async def main() -> None:
    parser = argparse.ArgumentParser(description="Ingestion catalogue Boss → PostgreSQL + pgvector")
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument(
        "--xlsx",
        type=Path,
        help="Fichier Excel BOSS (articles_industriels_1000.xlsx)",
    )
    source.add_argument(
        "--csv",
        type=Path,
        help="CSV legacy : sku,name,description,category,unit_price",
    )
    parser.add_argument(
        "--batch-pause",
        type=float,
        default=0.05,
        help="Pause entre chaque embedding (secondes, défaut 0.05)",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Limiter le nombre de lignes ingérées (0 = toutes)",
    )
    parser.add_argument(
        "--start-at",
        type=int,
        default=1,
        help="Reprendre à la ligne N (1-indexé)",
    )
    parser.add_argument(
        "--parse-only",
        action="store_true",
        help="Valider la lecture du fichier sans appeler la base ni l'API embedding",
    )
    args = parser.parse_args()

    if args.xlsx:
        if not args.xlsx.exists():
            raise SystemExit(f"Fichier introuvable : {args.xlsx}")
        rows = load_catalog_rows_from_xlsx(args.xlsx)
    else:
        if not args.csv.exists():
            raise SystemExit(f"Fichier introuvable : {args.csv}")
        rows = load_catalog_rows_from_csv(args.csv)

    if args.limit > 0:
        rows = rows[: args.limit]

    _print_parse_preview(rows)

    if args.parse_only:
        print("Mode --parse-only : aucune écriture en base.")
        return

    print(f"Démarrage ingestion de {len(rows)} articles (embeddings + PostgreSQL)…")
    await ingest_catalog(rows, batch_pause=args.batch_pause, start_at=args.start_at)
    print("Ingestion terminée.")


if __name__ == "__main__":
    asyncio.run(main())
